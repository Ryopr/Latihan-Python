import pandas as pd
import numpy as np

file_excel=pd.read_excel('D:/Gdrive/2026/training/Latihan Python/data tender/Data_Tender_LPSE KABUPATEN KUTAI BARAT 2012.xlsx', usecols='A');

# './data tender/Data_Tender_LPSE KABUPATEN KUTAI BARAT 2012.xlsx'
# print(file_excel.head())

matrix=file_excel.to_numpy()
data=matrix
# print('===================================')
# print('ini cara mengambil 1 kolom')
# print(data)#mengambil 1 kolom

file_excel2=pd.read_excel('D:/Gdrive/2026/training/Latihan Python/data tender/Data_Tender_LPSE KABUPATEN KUTAI BARAT 2012.xlsx', header=None);
# print('===================================')
# print('ini cara mengambil 1 baris')
matrix2=file_excel2.to_numpy()
# print(matrix2[10])#mengambil 1 baris
# print(len(matrix2[10]))#mengambil 1 baris
from pathlib import Path
from openpyxl import load_workbook
file_name_existing = 'contoh data ouput.xlsx'
sheet_name = 'Sheet1'

folder_path = Path(r"d:\Gdrive\2026\training\Latihan Python\data tender")# nama alamat folder excel yang terpisah2
filenames = [f.name for f in folder_path.iterdir() if f.is_file()] # pemanggilan sejumlah nama file dari folder terpisah2
# print(filenames);

for i in range(len(filenames)):
    wb = load_workbook(file_name_existing)
    current_max_row = wb[sheet_name].max_row
    
    print('proses copy dan append file '+filenames[i]+' ke file '+ file_name_existing);
    file_excel3=pd.read_excel('D:/Gdrive/2026/training/Latihan Python/data tender/'+filenames[i], header=None);# membaca alamat tiap file excel yang terpisah2

    data_bacaan=file_excel3.to_numpy()[1:] # pengambilan data 1 baris di bawah header sampai 
    data_export=pd.DataFrame(data_bacaan)
    # data_export.to_excel(file_name_existing, index=False, header=False)
    with pd.ExcelWriter(file_name_existing, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        data_export.to_excel(
        writer, 
        sheet_name=sheet_name, 
        startrow=current_max_row,  # Starts writing immediately after existing data
        index=False,               # Removes 0, 1, 2... row numbers
        header=False               # Removes top header text row
    )
    # print('jumlah baris data adalah')
    # print(len(file_excel3.to_numpy()))
